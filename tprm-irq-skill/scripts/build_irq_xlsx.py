#!/usr/bin/env python3
"""Render an Inherent Risk Questionnaire (IRQ) as a fillable Excel workbook.

Reads a question definition (default: irq_questions.json beside this script),
optionally filters the conditional Section C modules to a set of in-scope
domains, and writes a styled .xlsx with three sheets:

  - "Questionnaire" : the fillable form - dropdown answers + branching notes
  - "Scoring"       : live formulas that compute dimension scores, the overall
                      tier (with override escalation), and due-diligence triggers
  - "Lists"/"Key"   : hidden helper sheets backing the dropdowns and scoring

Because answers and scores are wired with formulas, the tier updates live in
Excel as the business owner fills the form.

Usage:
  python build_irq_xlsx.py
  python build_irq_xlsx.py --domains "Cybersecurity,Data Privacy,Operational Resilience"
  python build_irq_xlsx.py --domains "Cybersecurity" --supplier "Acme Ltd" --out Acme_IRQ.xlsx

Requires: openpyxl
"""
import argparse
import json
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))

# ---- styling -------------------------------------------------------------
NAVY = "1F3864"
LIGHT = "D9E1F2"
GREY = "F2F2F2"
AMBER = "FFF2CC"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def title_font(sz=14, color=WHITE):
    return Font(bold=True, size=sz, color=color)


# ---- selection -----------------------------------------------------------
def select_questions(spec, domains):
    """Return the ordered list of included questions for the given scope."""
    scope = {d.strip() for d in domains if d.strip()} if domains else None
    included = []
    for section in spec["sections"]:
        for q in section["questions"]:
            qdoms = set(q.get("domains", []))
            if "core" in qdoms:
                included.append((section, q))
            elif scope is None or (qdoms & scope):
                included.append((section, q))
    return included


# ---- build ---------------------------------------------------------------
def build(spec, domains, supplier, out_path):
    included = select_questions(spec, domains)
    inc_ids = [q["id"] for _, q in included]
    if len(inc_ids) > 30:
        raise SystemExit(f"Refusing to build: {len(inc_ids)} questions exceeds the 30 cap.")

    wb = Workbook()

    # ----- hidden Lists sheet (dropdown sources) --------------------------
    lists = wb.active
    lists.title = "Lists"
    list_range = {}  # qid -> "Lists!$C$2:$C$5"
    col = 1
    for _, q in included:
        if q["type"] != "choice":
            continue
        letter = lists.cell(row=1, column=col).column_letter
        lists.cell(row=1, column=col, value=q["id"])
        for i, opt in enumerate(q["options"], start=2):
            lists.cell(row=i, column=col, value=opt["label"])
        n = len(q["options"]) + 1
        list_range[q["id"]] = f"Lists!${letter}$2:${letter}${n}"
        col += 1
    lists.sheet_state = "hidden"

    # ----- hidden Key sheet (option -> score) -----------------------------
    key = wb.create_sheet("Key")
    key.cell(row=1, column=1, value="QID")
    key.cell(row=1, column=2, value="Option")
    key.cell(row=1, column=3, value="Score")
    kr = 2
    for _, q in included:
        for opt in q.get("options", []):
            if opt.get("score") is None:
                continue
            key.cell(row=kr, column=1, value=q["id"])
            key.cell(row=kr, column=2, value=opt["label"])
            key.cell(row=kr, column=3, value=opt["score"])
            kr += 1
    key.sheet_state = "hidden"

    # ----- Questionnaire sheet -------------------------------------------
    qs = wb.create_sheet("Questionnaire", 0)
    qs.column_dimensions["A"].width = 7
    qs.column_dimensions["B"].width = 72
    qs.column_dimensions["C"].width = 34
    qs.column_dimensions["D"].width = 40

    qs.merge_cells("A1:D1")
    c = qs["A1"]
    c.value = spec["title"]
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.font = title_font(15)
    c.alignment = CENTER
    qs.row_dimensions[1].height = 26

    qs.merge_cells("A2:D2")
    intro = qs["A2"]
    intro.value = spec["intro"]
    intro.alignment = WRAP
    intro.fill = PatternFill("solid", fgColor=GREY)
    qs.row_dimensions[2].height = 70

    # supplier / respondent fields
    qs["A3"], qs["B3"] = ("Supplier:", supplier or "")
    qs["C3"] = "Completed by:"
    qs["A4"], qs["C4"] = ("Date:", "Risk tier (auto):")
    qs["D4"] = "=Scoring!$B$2"
    for ref in ("A3", "C3", "A4", "C4"):
        qs[ref].font = Font(bold=True)
    qs["D4"].font = Font(bold=True, color=NAVY)

    answer_cell = {}  # qid -> "Questionnaire!$C$row"
    row = 6
    current_section = None
    for section, q in included:
        if section["id"] != current_section:
            current_section = section["id"]
            qs.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            hdr = qs.cell(row=row, column=1, value=f"{section['name']}")
            hdr.fill = PatternFill("solid", fgColor=LIGHT)
            hdr.font = Font(bold=True, size=12, color=NAVY)
            hdr.alignment = Alignment(vertical="center")
            qs.row_dimensions[row].height = 20
            row += 1
            # column captions
            for col_i, cap in ((1, "Q#"), (2, "Question"), (3, "Your answer"), (4, "Guidance")):
                cc = qs.cell(row=row, column=col_i, value=cap)
                cc.font = Font(bold=True, italic=True, size=9, color="808080")
            row += 1

        qs.cell(row=row, column=1, value=q["id"]).alignment = WRAP
        qtext = qs.cell(row=row, column=2, value=q["text"])
        qtext.alignment = WRAP
        ans = qs.cell(row=row, column=3)
        ans.fill = PatternFill("solid", fgColor=AMBER)
        ans.border = BORDER
        ans.alignment = WRAP

        guidance = []
        if q.get("gate"):
            guidance.append(f"Only if {q['gate']}")
        if q.get("branch_note"):
            guidance.append(q["branch_note"])
        if guidance:
            g = qs.cell(row=row, column=4, value="  ".join(guidance))
            g.alignment = WRAP
            g.font = Font(size=9, color="808080")

        if q["type"] == "choice" and q["id"] in list_range:
            dv = DataValidation(type="list", formula1=list_range[q["id"]], allow_blank=True)
            qs.add_data_validation(dv)
            dv.add(ans)
        answer_cell[q["id"]] = f"Questionnaire!$C${row}"
        qs.row_dimensions[row].height = 30
        row += 1

    # ----- Scoring sheet --------------------------------------------------
    sc = wb.create_sheet("Scoring")
    sc.column_dimensions["A"].width = 30
    sc.column_dimensions["B"].width = 22
    sc.column_dimensions["C"].width = 10
    sc.column_dimensions["D"].width = 52

    sc.merge_cells("A1:D1")
    h = sc["A1"]
    h.value = "Scoring (internal - not shown to the respondent)"
    h.fill = PatternFill("solid", fgColor=NAVY)
    h.font = title_font(12)
    sc["A2"] = "FINAL RISK TIER"
    sc["A2"].font = Font(bold=True)
    sc["B2"] = None  # set after final_rank computed (formula below)
    sc["B2"].font = Font(bold=True, size=12, color=NAVY)

    # per-question score block
    r = 4
    sc.cell(row=r, column=1, value="Per-question score").font = Font(bold=True)
    r += 1
    head = ["Question", "Answer", "Score"]
    for i, hh in enumerate(head):
        sc.cell(row=r, column=1 + i, value=hh).font = Font(bold=True, italic=True, size=9, color="808080")
    r += 1
    score_cell = {}  # qid -> "C{row}"
    for _, q in included:
        if all(o.get("score") is None for o in q.get("options", [])):
            continue  # unscored (A1/A2)
        sc.cell(row=r, column=1, value=q["id"])
        sc.cell(row=r, column=2, value=f"={answer_cell[q['id']]}").alignment = WRAP
        sc.cell(
            row=r, column=3,
            value=f"=SUMIFS(Key!$C:$C,Key!$A:$A,\"{q['id']}\",Key!$B:$B,{answer_cell[q['id']]})",
        )
        score_cell[q["id"]] = f"C{r}"
        r += 1

    # dimension rollup
    r += 1
    dim_start = r
    sc.cell(row=r, column=1, value="Dimension scores (max of inputs)").font = Font(bold=True)
    r += 1
    dim_score_rows = {}
    for dim in spec["dimensions"]:
        members = [score_cell[qid] for qid in spec["dimension_map"][dim] if qid in score_cell]
        sc.cell(row=r, column=1, value=spec["dimension_labels"][dim])
        if members:
            sc.cell(row=r, column=2, value="=MAX(" + ",".join(members) + ")")
        else:
            sc.cell(row=r, column=2, value=0)
        dim_score_rows[dim] = r
        r += 1
    dim_first = dim_start + 1
    dim_last = r - 1
    dim_range = f"B{dim_first}:B{dim_last}"

    # average / computed tier
    r += 1
    sc.cell(row=r, column=1, value="Average of applicable dimensions").font = Font(bold=True)
    avg_cell = f"B{r}"
    sc.cell(row=r, column=2,
            value=f'=IFERROR(SUMIF({dim_range},">0")/COUNTIF({dim_range},">0"),0)')
    r += 1
    sc.cell(row=r, column=1, value="Computed tier rank")
    comp_rank_cell = f"B{r}"
    sc.cell(row=r, column=2,
            value=f'=IF({avg_cell}>=3.5,4,IF({avg_cell}>=2.5,3,IF({avg_cell}>=1.5,2,IF({avg_cell}>0,1,0))))')
    r += 1

    # overrides
    r += 1
    sc.cell(row=r, column=1, value="Override checks").font = Font(bold=True)
    r += 1
    override_cells = []
    for ov in spec.get("overrides", []):
        refs = set(re.findall(r"[AB]\d|C-[A-Z0-9]+", ov["when"]))
        if not refs.issubset(set(inc_ids)):
            continue  # referenced question not in this scope
        formula = _override_formula(ov["when"], ov["min_rank"], answer_cell, score_cell)
        if formula is None:
            continue
        sc.cell(row=r, column=1, value=ov["id"])
        sc.cell(row=r, column=2, value=formula)
        sc.cell(row=r, column=4, value=ov["note"]).font = Font(size=9, color="808080")
        sc.cell(row=r, column=4).alignment = WRAP
        override_cells.append(f"B{r}")
        r += 1

    # final rank + tier text
    r += 1
    sc.cell(row=r, column=1, value="Final tier rank").font = Font(bold=True)
    final_rank_cell = f"B{r}"
    rank_inputs = [comp_rank_cell] + override_cells
    sc.cell(row=r, column=2, value="=MAX(" + ",".join(rank_inputs) + ")")
    r += 1
    # wire the headline B2 to the tier text
    sc["B2"] = (f'=IF({final_rank_cell}=4,"Critical",IF({final_rank_cell}=3,"High",'
                f'IF({final_rank_cell}=2,"Medium",IF({final_rank_cell}=1,"Low","-"))))')

    # due-diligence triggers
    r += 2
    sc.cell(row=r, column=1, value="Due-diligence triggered").font = Font(bold=True)
    r += 1
    for label, formula in _triggers(answer_cell, score_cell, inc_ids):
        sc.cell(row=r, column=1, value=label)
        sc.cell(row=r, column=2, value=formula)
        r += 1

    wb.save(out_path)
    return inc_ids


def _atom(token, answer_cell, score_cell):
    """Translate 'QID=Yes' / 'QID>=3' / 'QID=4' into an Excel boolean fragment."""
    m = re.match(r"\s*([AB]\d|C-[A-Z0-9]+)\s*(>=|<=|=)\s*(.+?)\s*$", token)
    if not m:
        return None
    qid, op, rhs = m.group(1), m.group(2), m.group(3)
    if re.fullmatch(r"-?\d+(\.\d+)?", rhs):  # numeric -> compare the score
        cell = score_cell.get(qid)
        if cell is None:
            return None
        return f"{cell}{op}{rhs}"
    # text (Yes/No/Don't know, possibly slash-separated) -> match the answer
    if qid not in answer_cell:
        return None
    vals = [v.strip() for v in rhs.split("/")]
    parts = [f'{answer_cell[qid]}="{v}"' for v in vals]
    return parts[0] if len(parts) == 1 else "OR(" + ",".join(parts) + ")"


def _override_formula(when, min_rank, answer_cell, score_cell):
    if " AND " in when:
        atoms = [_atom(a, answer_cell, score_cell) for a in when.split(" AND ")]
        if any(a is None for a in atoms):
            return None
        cond = "AND(" + ",".join(atoms) + ")"
    elif " OR " in when:
        atoms = [_atom(a, answer_cell, score_cell) for a in when.split(" OR ")]
        if any(a is None for a in atoms):
            return None
        cond = "OR(" + ",".join(atoms) + ")"
    else:
        cond = _atom(when, answer_cell, score_cell)
        if cond is None:
            return None
    return f"=IF({cond},{min_rank},0)"


def _triggers(answer_cell, score_cell, inc_ids):
    """Return (label, formula) pairs for triggers whose questions are in scope."""
    inc = set(inc_ids)
    A = answer_cell
    S = score_cell
    out = []

    def add(label, need, cond_fn):
        # cond_fn is deferred so out-of-scope question ids never get looked up
        if set(need).issubset(inc):
            out.append((label, f'=IF({cond_fn()},"Yes - trigger","")'))

    add("Data Privacy DD + DPA / transfer", ["B1", "C-P1", "C-P2"],
        lambda: f'AND({A["B1"]}="Yes",OR({A["C-P2"]}="Yes",{S["C-P1"]}>=3))')
    add("Cybersecurity due-diligence questionnaire", ["B3", "C-IT1"],
        lambda: f'AND({A["B3"]}="Yes",{S["C-IT1"]}>=3)')
    add("Operational resilience / BCP-DR review", ["B4", "C-O1", "C-O2"],
        lambda: f'AND({A["B4"]}="Yes",OR({S["C-O1"]}>=3,{S["C-O2"]}>=3))')
    add("ABAC + sanctions / adverse-media screening", ["B5", "C-A1", "C-A2"],
        lambda: f'AND({A["B5"]}="Yes",OR({A["C-A1"]}="Yes",{A["C-A2"]}="Yes"))')
    add("Fourth-party / sub-processor disclosure", ["B6"],
        lambda: f'OR({A["B6"]}="Yes",{A["B6"]}="Don\'t know")')
    add("AI / model-risk assessment", ["B8", "C-AI1"],
        lambda: f'AND({A["B8"]}="Yes",{A["C-AI1"]}="Yes")')
    add("Cross-border transfer assessment", ["B9", "C-G1"],
        lambda: f'AND(OR({A["B9"]}="Yes",{A["B9"]}="Don\'t know"),{S["C-G1"]}>=3)')
    add("ESG / supply-chain due diligence", ["C-E1"], lambda: f'{A["C-E1"]}="Yes"')
    add("Financial viability assessment", ["C-FIN1"], lambda: f'{A["C-FIN1"]}="Yes"')
    return out


def main():
    ap = argparse.ArgumentParser(description="Render an IRQ as a fillable Excel workbook.")
    ap.add_argument("--questions", default=os.path.join(HERE, "irq_questions.json"))
    ap.add_argument("--domains", default="", help="Comma-separated in-scope domains; empty = all 30.")
    ap.add_argument("--supplier", default="")
    ap.add_argument("--out", default="IRQ.xlsx")
    args = ap.parse_args()

    with open(args.questions, encoding="utf-8") as f:
        spec = json.load(f)

    domains = [d for d in args.domains.split(",")] if args.domains else []
    ids = build(spec, domains, args.supplier, args.out)
    scope_desc = args.domains if args.domains else "all domains"
    print(f"Wrote {args.out} - {len(ids)} questions (scope: {scope_desc}).")
    print("Questions:", ", ".join(ids))


if __name__ == "__main__":
    main()
